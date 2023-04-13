import openpyxl

# select active sheet
book = openpyxl.load_workbook("C:\\Users\\User\\Documents\\PythonDemo.xlsx")
sheet = book.active

data_dict = {}

# get cell from the sheet
cell = sheet.cell(row=1, column=2)

#extract value presented in that box
print(cell.value)

# give value to Testcase 1 firstname
firstname1 = sheet.cell(row=2, column=2).value = "Neil"

print(firstname1)

# if you want to know total maximum rows present in the sheet
print("{} row/s present".format(sheet.max_row))

# if you want to know total maximum column present in the sheet
print("{} column/s present".format(sheet.max_column))

# prints Testcase4
print(sheet['A5'].value)

print("-------------------------------")

# prints all the values present in the excel sheet
for rows in range(1, sheet.max_row+1): # to get rows
    if sheet.cell(row=rows, column=1).value == "Testcase2":  #print only specific data

        for columns in range(2, sheet.max_column+1):   # to get columns
            #data_dict["email"] = "helloworld@gmail.com"
            data_dict[sheet.cell(row=1, column=columns).value] = sheet.cell(row=rows, column=columns).value

print(data_dict)