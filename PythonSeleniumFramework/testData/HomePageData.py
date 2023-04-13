#Home page test data should be placed here in this file
import openpyxl


class HomePageData:
    test_home_page_data = [{"name": "Neil", "email": "helloword@gmail.com", "password": "helloworld", "gender": "Male"},
                           {"name": "Charry", "email": "hiworld@gmail.com", "password": "hiworld", "gender": "Female"}]

    @staticmethod
    def get_test_data(test_case_name):   # no need to put self if @staticmethod is present
        #book = openpyxl.load_workbook("C:\\Users\\User\\Documents\\PythonDemo.xlsx")
        book = openpyxl.load_workbook("C:\\Users\\User\\PycharmProjects1\\PythonSeleniumFramework\\PythonDemo.xlsx")
        sheet = book.active

        data_dict = {}
        # prints all the values present in the excel sheet
        for rows in range(1, sheet.max_row + 1):  # to get rows
            if sheet.cell(row=rows, column=1).value == test_case_name:  # print only specific data

                for columns in range(2, sheet.max_column + 1):  # to get columns
                    # data_dict["email"] = "helloworld@gmail.com"
                    data_dict[sheet.cell(row=1, column=columns).value] = sheet.cell(row=rows, column=columns).value

        return [data_dict]